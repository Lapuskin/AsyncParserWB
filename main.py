# from datetime import time
# from random import uniform
# from time import sleep

import requests
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# def add_random_delay():
#     delay = uniform(0, 1)
#     sleep(delay)
#
#
# def get_filters(query, shard):
# '''
#     получает фильтры категории товара, путем запроса на сервер.
# '''
#     filters_url = 'https://catalog.wb.ru/catalog/' + shard + '/v4/filters?appType=1&' + query + '&curr=rub&dest=-59202&spp=30'
#     add_random_delay()
#     try:
#         filters_json = requests.get(filters_url).json()['data']['filters']
#         for filter in filters_json:
#             pass
#         print('чекнули фильтры')
#         return str(filters_json)
#     except Exception as e:
#         print(e, locals())
#         print()
#     return


def recursive_search(categories, wb: Workbook, ws: Worksheet=None,  row_column: list = [1, 1], level: int = 0,  is_main_category=False, ):
    for category in categories:
        print(category['name'])
        if is_main_category:
            ws = wb.create_sheet(category['name'])
            row_column = [1, 1]
        if 'query' in category:
            ws.cell(row=row_column[0], column=row_column[1], value=category['id'])
            ws.cell(row=row_column[0], column=row_column[1] + 1, value=category['name'])
            ws.cell(row=row_column[0], column=row_column[1] + 2, value=level)
            # далее код для парсинга всех фильтров категории товаров, который, при необходимости, можно будет использовать.
            #     filters = get_filters(category['query'], category['shard'])
            #     ws.cell(row=row_column[0], column=row_column[1] + 3, value=filters)
            row_column[0] += 1
            if 'childs' in category:
                row_column = recursive_search(categories=category['childs'], ws=ws, row_column=row_column, level=level + 1, wb=wb)
    return row_column


def main():
    menu_url = 'https://static-basket-01.wbbasket.ru/vol0/data/main-menu-by-ru-v2.json'
    cats_json = requests.get(menu_url).json()
    wb = Workbook()
    recursive_search(categories=cats_json, is_main_category=True, wb=wb)
    wb.save('test.xlsx')


if __name__ == '__main__':
    main()



